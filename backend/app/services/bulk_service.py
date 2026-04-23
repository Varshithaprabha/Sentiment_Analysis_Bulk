"""
services/bulk_service.py
========================
Handles CSV/Excel file processing: load → classify → aggregate → export.
"""

import pandas as pd
import json
import tempfile
import shutil
from pathlib import Path
from collections import Counter
from openpyxl.styles import PatternFill, Font

from app.core.sentiment_engine import classify_batch
from app.core.theme_extractor import extract_themes

FEEDBACK_COLS = ["feedback","comment","comments","text","response","review","description","message","opinion"]
CATEGORY_COLS = ["category","department","course","faculty","subject","type","course_name","module","instructor"]

COLORS = {"POSITIVE": "C8F7C5", "NEGATIVE": "F9CBCB", "NEUTRAL": "FFF4CC"}


def _detect_col(df, candidates):
    lower_map = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c in lower_map:
            return lower_map[c]
    return None


def process_file(filepath: str) -> dict:
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        # Dynamic decoding & delimiter detection
        for enc in ["utf-8", "latin-1", "utf-16", "cp1252"]:
            try:
                df = pd.read_csv(filepath, encoding=enc, sep=None, engine="python")
                break
            except Exception:
                continue
        else:
            raise ValueError("Could not decode CSV with supported encodings (UTF-8, Latin-1, cp1252).")
    elif suffix in (".xlsx", ".xls"):
        df = pd.read_excel(filepath, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # Remove completely empty rows
    df = df.dropna(how="all").reset_index(drop=True)

    feedback_col = _detect_col(df, FEEDBACK_COLS)
    
    # DYNAMIC LOGIC: If no standard header exists, find the column with the highest avg text length
    if not feedback_col:
        str_cols = df.select_dtypes(include=["object", "string"]).columns.tolist()
        if not str_cols:
            raise ValueError("No text-based columns found in the dataset.")
        
        # Calculate mean word count to find the most "comment-like" column
        lengths = {col: df[col].astype(str).str.len().mean() for col in str_cols}
        feedback_col = max(lengths, key=lengths.get)

    category_col = _detect_col(df, CATEGORY_COLS)
    
    # DYNAMIC LOGIC: If no category header exists, find a potential categorical column 
    # (Low unique count relative to total rows, but > 1)
    if not category_col:
        potential_cats = []
        for col in df.columns:
            if col == feedback_col: continue
            unique_count = df[col].nunique()
            if 1 < unique_count < (len(df) * 0.4): # Less than 40% unique suggests categories
                potential_cats.append(col)
        if potential_cats:
            # Pick the one with the most balanced distribution or just the first
            category_col = potential_cats[0]

    texts = df[feedback_col].fillna("").astype(str).tolist()
    results = classify_batch(texts)

    df = df.copy()
    df["sentiment"]  = [r["sentiment"]  for r in results]
    df["sentiment_icon"] = [r["emoji"] for r in results]
    df["confidence"] = [r["confidence"] for r in results]
    df["pos_score"]  = [r["scores"]["positive"] for r in results]
    df["neg_score"]  = [r["scores"]["negative"] for r in results]

    # Summary stats
    total = len(df)
    counts = Counter(df["sentiment"])

    summary = {
        "total_responses": total,
        "sentiment_distribution": {
            lbl: {"count": counts.get(lbl, 0),
                  "percentage": round(100 * counts.get(lbl, 0) / max(total, 1), 1)}
            for lbl in ["POSITIVE", "NEGATIVE", "NEUTRAL"]
        },
        "average_confidence": round(float(df["confidence"].mean()), 4),
        "category_breakdown": [],
        "themes": extract_themes(results),
        "feedback_col": feedback_col,
        "category_col": category_col
    }

    # Category breakdown
    if category_col and category_col in df.columns:
        for cat, grp in df.groupby(category_col):
            cc = Counter(grp["sentiment"])
            ct = len(grp)
            dominant = max(cc, key=cc.get) if cc else "NEUTRAL"
            summary["category_breakdown"].append({
                "category": str(cat),
                "total": ct,
                "positive_pct": round(100 * cc.get("POSITIVE", 0) / ct, 1),
                "negative_pct": round(100 * cc.get("NEGATIVE", 0) / ct, 1),
                "neutral_pct":  round(100 * cc.get("NEUTRAL", 0) / ct, 1),
                "avg_confidence": round(float(grp["confidence"].mean()), 4),
                "dominant_sentiment": dominant
            })

    # Save Excel to temp dir
    tmp = tempfile.mkdtemp()
    excel_path = Path(tmp) / "detailed_results.xlsx"
    csv_path = Path(tmp) / "detailed_results.csv"
    
    # 1. Excel with Formatting
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Full Analysis")
        
        # Add a Summary sheet if categories exist
        if summary["category_breakdown"]:
            cat_df = pd.DataFrame(summary["category_breakdown"])
            cat_df.to_excel(writer, index=False, sheet_name="Category Summary")
            
        ws = writer.sheets["Full Analysis"]
        headers = [cell.value for cell in ws[1]]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")

        sent_idx = headers.index("sentiment") + 1 if "sentiment" in headers else None
        if sent_idx:
            for row in ws.iter_rows(min_row=2):
                val = row[sent_idx - 1].value
                fill = PatternFill("solid", fgColor=COLORS.get(val, "FFFFFF"))
                for cell in row: cell.fill = fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 60)

    # 2. Simple CSV Export
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    excel_bytes = excel_path.read_bytes()
    csv_bytes = csv_path.read_bytes()
    shutil.rmtree(tmp, ignore_errors=True)

    return {"summary": summary, "excel_bytes": excel_bytes, "csv_bytes": csv_bytes}
